"""
report_engine.py
-----------------
Core logic for the Returns Reconciliation tool.

Cross-checks marketplace return/refund exports (Shopee, Lazada, TikTok) against an
existing brand "Return Report" tracker workbook, finds return Order IDs that are not
yet tracked, and builds the rows that should be appended -- enriching missing fields
(Invoice Number, SKU) from a TC Order Report when available.

This module has NO Streamlit dependency so it can be imported, unit tested, or reused
in a CLI / notebook context independently of the app.
"""

from __future__ import annotations

import copy
import io
from dataclasses import dataclass, field
from datetime import datetime
from typing import Any, Optional

import openpyxl
import pandas as pd
from openpyxl.worksheet.worksheet import Worksheet


# ---------------------------------------------------------------------------
# Data structures
# ---------------------------------------------------------------------------

@dataclass
class NewReturnRow:
    """A single new return row ready to be appended to the tracker."""
    order_id: str
    platform: str
    return_request_date: Optional[datetime] = None
    ordered_date: Optional[datetime] = None
    invoice_number: Optional[str] = None
    tracking_number: Optional[str] = None
    sku: Optional[str] = None
    qty: Optional[float] = None
    return_reason: Optional[str] = None
    source: str = ""  # which marketplace file this came from
    notes: str = ""   # flags e.g. "no TC match found"


@dataclass
class ReconciliationResult:
    """Summary + detail of a reconciliation run for one marketplace."""
    marketplace: str
    total_in_file: int = 0
    already_tracked: list[str] = field(default_factory=list)
    new_rows: list[NewReturnRow] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)


# ---------------------------------------------------------------------------
# ID normalization
# ---------------------------------------------------------------------------

def normalize_order_id(value: Any) -> str:
    """
    Normalize an order ID for set comparison.

    Numeric marketplace IDs (notably Lazada) are frequently loaded as floats by
    openpyxl/pandas (e.g. 167605744763442.0) while elsewhere they appear as plain
    strings (167605744763442). Without normalization these compare as different
    and produce false "new" positives.
    """
    if value is None:
        return ""
    if isinstance(value, float):
        # Guard against genuine non-integer floats (shouldn't happen for IDs, but
        # safer than blindly truncating).
        if value.is_integer():
            return str(int(value))
        return str(value)
    return str(value).strip()


# ---------------------------------------------------------------------------
# Tracker workbook helpers
# ---------------------------------------------------------------------------

# Header names we look for to confirm we found the real header row inside the
# tracker's main sheet (title/instruction rows usually precede it).
_HEADER_PROBE_KEYWORDS = ("order number", "order no", "order id")


def find_main_sheet_name(wb: openpyxl.Workbook) -> str:
    """Find the sheet that looks like the main 'Return report' sheet."""
    candidates = [s for s in wb.sheetnames if "return report" in s.lower()]
    if candidates:
        return candidates[0]
    # Fallback: first sheet
    return wb.sheetnames[0]


def find_header_row(ws: Worksheet, max_scan_rows: int = 15) -> int:
    """
    Scan the first `max_scan_rows` rows to find the real header row -- the one
    containing something like 'Order Number'. Returns a 0-indexed row position
    within ws.iter_rows() output (NOT the Excel row number).
    """
    rows = list(ws.iter_rows(values_only=True, max_row=max_scan_rows))
    for i, row in enumerate(rows):
        text_cells = [str(c).strip().lower() for c in row if c is not None]
        if any(any(k in cell for k in _HEADER_PROBE_KEYWORDS) for cell in text_cells):
            return i
    # Fallback to the convention seen in existing brand trackers
    return 4


def build_header_index(header_row: tuple) -> dict[str, int]:
    """Map lowercased header text -> column index (0-based)."""
    idx = {}
    for i, h in enumerate(header_row):
        if h is not None:
            idx[str(h).strip().lower()] = i
    return idx


def _find_col(header_index: dict[str, int], *candidates: str) -> Optional[int]:
    """Return the first matching column index for any of the candidate header names."""
    for c in candidates:
        if c in header_index:
            return header_index[c]
    # loose contains-match fallback
    for key, i in header_index.items():
        for c in candidates:
            if c in key:
                return i
    return None


def load_tracker(file_like) -> dict:
    """
    Load the tracker workbook and return everything downstream steps need:
      - workbook object (kept open, editable)
      - main sheet name + worksheet
      - header row index (0-based within sheet) and Excel row number
      - header_index: lowercased header name -> column index
      - existing_order_ids: normalized set of Order Numbers already present
      - last_data_row: last Excel row number containing data (for appending)
    """
    wb = openpyxl.load_workbook(file_like, data_only=False)
    sheet_name = find_main_sheet_name(wb)
    ws = wb[sheet_name]

    header_row_idx0 = find_header_row(ws)
    rows = list(ws.iter_rows(values_only=True))
    header_row = rows[header_row_idx0]
    header_index = build_header_index(header_row)

    order_col = _find_col(header_index, "order number", "order no", "order id")
    if order_col is None:
        raise ValueError(
            f"Could not find an Order Number column in sheet '{sheet_name}'. "
            f"Headers found: {list(header_index.keys())}"
        )

    existing_order_ids: set[str] = set()
    last_data_row_excel = header_row_idx0 + 1  # 1-indexed Excel row of header
    for offset, row in enumerate(rows[header_row_idx0 + 1:]):
        excel_row = header_row_idx0 + 2 + offset  # 1-indexed
        if any(v is not None for v in row):
            last_data_row_excel = excel_row
            val = row[order_col] if order_col < len(row) else None
            norm = normalize_order_id(val)
            if norm:
                existing_order_ids.add(norm)

    return {
        "workbook": wb,
        "sheet_name": sheet_name,
        "worksheet": ws,
        "header_row_idx0": header_row_idx0,
        "header_row_excel": header_row_idx0 + 1,
        "header": header_row,
        "header_index": header_index,
        "order_col": order_col,
        "existing_order_ids": existing_order_ids,
        "last_data_row_excel": last_data_row_excel,
    }


# ---------------------------------------------------------------------------
# Marketplace file readers
# ---------------------------------------------------------------------------

def read_shopee_returns(file_like) -> pd.DataFrame:
    """
    Shopee return/refund export. Handles both legacy .xls (read via the xlrd
    engine) and modern .xlsx. Expected columns include:
    'Return ID', 'Order ID', 'Order Creation Date', 'Product Name', 'SKU',
    'Return Quantity', 'Return Reason', 'Return Creation Time',
    'Return Tracking Number', 'Return / Refund Status'.
    """
    df = _read_excel_any(file_like)
    df.columns = [str(c).strip() for c in df.columns]
    if "Order ID" not in df.columns:
        raise ValueError(f"Shopee file missing 'Order ID' column. Found: {list(df.columns)}")
    df["Order ID"] = df["Order ID"].apply(normalize_order_id)
    return df


def read_lazada_returns(file_like) -> pd.DataFrame:
    """
    Lazada return export. Expected columns include 'Order ID', 'Return Order ID',
    'Order Date', 'Return Order Date', 'Seller SKU ID' (naming can vary slightly).
    """
    df = _read_excel_any(file_like)
    df.columns = [str(c).strip() for c in df.columns]
    if "Order ID" not in df.columns:
        raise ValueError(f"Lazada file missing 'Order ID' column. Found: {list(df.columns)}")
    df["Order ID"] = df["Order ID"].apply(normalize_order_id)
    return df


def read_tiktok_returns(file_like) -> pd.DataFrame:
    """
    TikTok return export.

    TikTok's xlsx export has a structural defect: every cell is wrapped in its
    own <row r="1"> element in the sheet XML, so all rows share r=1. This means
    openpyxl read_only mode (and pandas via openpyxl) sees only ONE row -- the
    header. We work around this by parsing the ZIP/XML directly, reconstructing
    the true row layout from each cell's column-letter+row-number reference
    (e.g. "R5" = column R, row 5).

    Active statuses to keep: "In Process", "To Process".
    Completed / Refund rejected rows are excluded.

    The function returns a DataFrame with a standardised "Order ID" column (using
    "Return Order ID" as the key, since that's what TikTok exports) plus all
    other columns from the export (SKU, Return Reason, Tracking, Return Status,
    Time Requested, Qty, etc.).
    """
    import zipfile
    import xml.etree.ElementTree as ET
    import re as _re
    from collections import defaultdict

    KEEP_STATUSES = {"In Process", "To Process"}
    _NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"

    # --- Read raw bytes ---
    if isinstance(file_like, (bytes, bytearray)):
        raw = bytes(file_like)
    elif hasattr(file_like, "read"):
        raw = file_like.read()
        if hasattr(file_like, "seek"):
            file_like.seek(0)
    else:
        with open(file_like, "rb") as fh:
            raw = fh.read()

    # --- Try standard read first (handles non-malformed TikTok files) ---
    try:
        df_try = pd.read_excel(io.BytesIO(raw), engine="openpyxl")
        if len(df_try) > 0 and len(df_try.columns) > 1:
            # File parsed correctly with multiple columns and rows
            df_try.columns = [str(c).strip() for c in df_try.columns]
            _standardise_tiktok_order_id(df_try)
            if "Return Status" in df_try.columns:
                df_try = df_try[df_try["Return Status"].isin(KEEP_STATUSES)]
            return df_try
    except Exception:
        pass

    # --- XML fallback for malformed TikTok exports ---
    with zipfile.ZipFile(io.BytesIO(raw)) as z:
        sheet_paths = [n for n in z.namelist() if n.startswith("xl/worksheets/") and n.endswith(".xml")]
        if not sheet_paths:
            raise ValueError("TikTok xlsx file contains no worksheet XML.")
        with z.open(sheet_paths[0]) as f:
            xml_content = f.read().decode("utf-8")

    root = ET.fromstring(xml_content)
    sd = root.find(f"{{{_NS}}}sheetData")
    if sd is None:
        raise ValueError("TikTok sheet XML has no <sheetData> element.")

    cells: dict[int, dict[str, str]] = defaultdict(dict)
    for row_elem in sd.findall(f"{{{_NS}}}row"):
        for c_elem in row_elem.findall(f"{{{_NS}}}c"):
            ref = c_elem.get("r", "")
            m = _re.match(r"([A-Z]+)(\d+)", ref)
            if m:
                col_letter, row_num = m.group(1), int(m.group(2))
                v = c_elem.find(f"{{{_NS}}}v")
                cells[row_num][col_letter] = v.text if v is not None else None

    if not cells:
        return pd.DataFrame(columns=["Order ID"])

    # Row 1 = header; rows 2+ = data
    col_letters = sorted(
        {letter for row_dict in cells.values() for letter in row_dict},
        key=lambda x: (len(x), x),
    )
    header_map = {letter: cells[1].get(letter) for letter in col_letters}
    col_names = [header_map.get(l, l) for l in col_letters]

    records = []
    for row_num in sorted(k for k in cells if k > 1):
        row_dict = cells[row_num]
        records.append({col_names[i]: row_dict.get(l) for i, l in enumerate(col_letters)})

    df = pd.DataFrame(records)
    df.columns = [str(c).strip() for c in df.columns]

    # Filter to active statuses only
    status_col = next((c for c in df.columns if c.lower() == "return status"), None)
    if status_col:
        df = df[df[status_col].isin(KEEP_STATUSES)].reset_index(drop=True)

    _standardise_tiktok_order_id(df)
    return df


def _standardise_tiktok_order_id(df: pd.DataFrame) -> None:
    """Ensure the DataFrame has a normalised 'Order ID' column in-place."""
    if "Order ID" in df.columns:
        df["Order ID"] = df["Order ID"].apply(normalize_order_id)
        return
    for c in df.columns:
        if "order id" in c.lower():
            df["Order ID"] = df[c].apply(normalize_order_id)
            return
    raise ValueError(f"TikTok file missing an Order ID column. Found: {list(df.columns)}")


def _read_excel_any(file_like) -> pd.DataFrame:
    """
    Read an Excel file regardless of whether it is a legacy .xls or modern .xlsx.
    Accepts a path string, raw bytes, or any file-like object (e.g. Streamlit's
    UploadedFile, io.BytesIO).

    Detection uses magic bytes so we never rely on the file extension, which is
    often absent or wrong when Streamlit wraps uploaded files in a BytesIO buffer.
      0xD0CF11E0... = OLE2 compound document -> legacy .xls  -> xlrd
      PK\x03\x04    = ZIP archive            -> modern .xlsx -> openpyxl
    """
    if isinstance(file_like, (bytes, bytearray)):
        raw = bytes(file_like)
    elif hasattr(file_like, "read"):
        raw = file_like.read()
        if hasattr(file_like, "seek"):
            file_like.seek(0)
    else:
        with open(file_like, "rb") as fh:
            raw = fh.read()

    is_xls  = raw[:8] == b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"
    is_xlsx = raw[:4] == b"PK\x03\x04"

    if is_xls:
        return pd.read_excel(io.BytesIO(raw), engine="xlrd")
    if is_xlsx:
        try:
            return pd.read_excel(io.BytesIO(raw), engine="openpyxl")
        except Exception:
            return pd.read_excel(io.BytesIO(raw), engine="xlrd")
    # inconclusive magic -- try both
    try:
        return pd.read_excel(io.BytesIO(raw), engine="openpyxl")
    except Exception:
        return pd.read_excel(io.BytesIO(raw), engine="xlrd")


def read_tc_order_report(file_like) -> pd.DataFrame:
    """
    TC Order Report CSV. Expected columns include 'order_id', 'invoice_number',
    'sku', 'custom_sku', 'order_status', 'item_title' (exact set can vary by
    client export, hence the dtype=str + strip pass).
    """
    df = pd.read_csv(file_like, dtype=str)
    df.columns = [c.strip() for c in df.columns]
    for c in df.columns:
        df[c] = df[c].astype(str).str.strip()
    if "order_id" not in df.columns:
        raise ValueError(f"TC Order Report missing 'order_id' column. Found: {list(df.columns)}")
    df["order_id"] = df["order_id"].apply(normalize_order_id)
    return df


# ---------------------------------------------------------------------------
# Reconciliation logic
# ---------------------------------------------------------------------------

def _safe_get(row: pd.Series, *names: str):
    for n in names:
        if n in row.index and pd.notna(row[n]):
            return row[n]
    return None


def reconcile_marketplace(
    marketplace: str,
    mp_df: pd.DataFrame,
    existing_order_ids: set[str],
    tc_df: Optional[pd.DataFrame] = None,
) -> ReconciliationResult:
    """
    Compare one marketplace return file's Order IDs against the tracker's existing
    set, and build NewReturnRow objects for anything not yet tracked -- enriched
    from the TC Order Report where possible.
    """
    result = ReconciliationResult(marketplace=marketplace)

    if mp_df.empty:
        result.warnings.append(
            f"{marketplace} file has no data rows (header only) -- nothing to reconcile."
        )
        return result

    if "Order ID" not in mp_df.columns:
        result.warnings.append(f"{marketplace} file has no recognizable Order ID column.")
        return result

    all_ids = mp_df["Order ID"].dropna().unique().tolist()
    result.total_in_file = len(all_ids)

    new_ids = [oid for oid in all_ids if oid not in existing_order_ids]
    result.already_tracked = [oid for oid in all_ids if oid in existing_order_ids]

    # Group rows by order id in case one order has multiple SKU line items
    grouped = mp_df[mp_df["Order ID"].isin(new_ids)].groupby("Order ID")

    for order_id, group in grouped:
        first = group.iloc[0]

        tc_match = pd.DataFrame()
        if tc_df is not None:
            tc_match = tc_df[tc_df["order_id"] == order_id]

        invoice = None
        sku = None
        if not tc_match.empty:
            invoice = tc_match.iloc[0].get("invoice_number")
            skus = [s for s in tc_match.get("sku", pd.Series(dtype=str)).tolist() if s and s != "nan"]
            sku = " / ".join(dict.fromkeys(skus)) if skus else None
            notes = ""
        else:
            notes = "no TC match found"

        # Fall back to marketplace file's own SKU/product info if TC didn't have it.
        # Different marketplaces name this column differently (Shopee: 'SKU',
        # TikTok: 'Seller SKU', Lazada: 'Seller SKU ID').
        if not sku:
            sku_col = None
            for candidate in ("SKU", "Seller SKU", "Seller SKU ID"):
                if candidate in group.columns:
                    sku_col = candidate
                    break
            if sku_col:
                mp_skus = [str(s) for s in group[sku_col].tolist() if pd.notna(s)]
                sku = " / ".join(dict.fromkeys(mp_skus)) if mp_skus else None
            if not sku:
                sku = _safe_get(first, "Product Name")

        if not invoice:
            invoice = _safe_get(first, "Invoice Number", "invoice_number")

        return_request_date = _safe_get(first, "Return Creation Time", "Return Order Date", "Time Requested")
        ordered_date = _safe_get(first, "Order Creation Date", "Order Date")
        tracking = _safe_get(first, "Return Tracking Number", "Return Logistic Tracking Number", "Return Logistics Tracking ID")
        reason = _safe_get(first, "Return Reason")
        qty = _safe_get(first, "Return Quantity")

        result.new_rows.append(
            NewReturnRow(
                order_id=order_id,
                platform=marketplace,
                return_request_date=_coerce_datetime(return_request_date),
                ordered_date=_coerce_datetime(ordered_date),
                invoice_number=invoice,
                tracking_number=tracking,
                sku=sku,
                qty=qty,
                return_reason=reason,
                source=marketplace,
                notes=notes,
            )
        )

    return result


def _coerce_datetime(value) -> Optional[datetime]:
    if value is None or value == "" or (isinstance(value, float) and pd.isna(value)):
        return None
    if isinstance(value, datetime):
        return value
    try:
        return pd.to_datetime(value).to_pydatetime()
    except Exception:
        return None


# ---------------------------------------------------------------------------
# Writing new rows back into the tracker workbook
# ---------------------------------------------------------------------------

# Columns that should NEVER be auto-filled -- these require manual confirmation
# by ops/warehouse/SEG/GRAAS staff per the tracker's own "ACTION by" row.
MANUAL_COLUMN_KEYWORDS = (
    "return confirmation",
    "return receiving date",
    "fault description",
    "dispute",
    "redressing instruction",
    "action status",
    "status",
)

# Maps our NewReturnRow fields -> possible header text fragments in the tracker.
FIELD_HEADER_CANDIDATES: dict[str, tuple[str, ...]] = {
    "platform": ("marketplace", "platform"),
    "return_request_date": ("return request date",),
    "order_id": ("order number", "order no", "order id"),
    "ordered_date": ("ordered date", "invoice date", "order date"),
    "invoice_number": ("invoice number", "invoice no"),
    "tracking_number": ("tracking number", "logistic tracking"),
    "sku": ("sku",),
    "qty": ("qty", "quantity"),
    "return_reason": ("return reason",),
}


def append_new_rows(tracker: dict, new_rows: list[NewReturnRow]) -> int:
    """
    Append new_rows to the tracker's main sheet, copying formatting from the last
    existing data row and only writing into columns we have a confident header
    match for. Manual-confirmation columns are left blank intentionally.

    Returns the number of rows appended.
    """
    ws: Worksheet = tracker["worksheet"]
    header_index = tracker["header_index"]
    last_row = tracker["last_data_row_excel"]
    ref_row = last_row if last_row > tracker["header_row_excel"] else tracker["header_row_excel"] + 1
    n_cols = len(tracker["header"])

    # Resolve field -> column number (1-indexed) once
    field_col: dict[str, int] = {}
    for field_name, candidates in FIELD_HEADER_CANDIDATES.items():
        col0 = _find_col(header_index, *candidates)
        if col0 is not None:
            field_col[field_name] = col0 + 1  # openpyxl is 1-indexed

    start_row = last_row + 1
    for offset, row in enumerate(new_rows):
        excel_row = start_row + offset

        # Copy styling from a representative existing row
        for col in range(1, n_cols + 1):
            src = ws.cell(row=ref_row, column=col)
            dst = ws.cell(row=excel_row, column=col)
            dst.font = copy.copy(src.font)
            dst.fill = copy.copy(src.fill)
            dst.border = copy.copy(src.border)
            dst.alignment = copy.copy(src.alignment)
            dst.number_format = src.number_format

        values = {
            "platform": row.platform,
            "return_request_date": row.return_request_date,
            "order_id": row.order_id,
            "ordered_date": row.ordered_date,
            "invoice_number": row.invoice_number,
            "tracking_number": row.tracking_number,
            "sku": row.sku,
            "qty": row.qty,
            "return_reason": row.return_reason,
        }
        for field_name, value in values.items():
            col = field_col.get(field_name)
            if col is not None:
                ws.cell(row=excel_row, column=col, value=value)

    return len(new_rows)


# ---------------------------------------------------------------------------
# High-level orchestration
# ---------------------------------------------------------------------------

def run_reconciliation(
    tracker_file,
    shopee_file=None,
    lazada_file=None,
    tiktok_file=None,
    tc_file=None,
) -> tuple[openpyxl.Workbook, list[ReconciliationResult]]:
    """
    Full pipeline: load tracker, load whichever marketplace files were provided,
    reconcile each against the tracker, append new rows, and return the updated
    workbook plus a list of per-marketplace results for reporting in the UI.
    """
    tracker = load_tracker(tracker_file)
    tc_df = read_tc_order_report(tc_file) if tc_file is not None else None

    results: list[ReconciliationResult] = []
    all_new_rows: list[NewReturnRow] = []

    if shopee_file is not None:
        mp_df = read_shopee_returns(shopee_file)
        res = reconcile_marketplace("Shopee", mp_df, tracker["existing_order_ids"], tc_df)
        results.append(res)
        all_new_rows.extend(res.new_rows)

    if lazada_file is not None:
        mp_df = read_lazada_returns(lazada_file)
        res = reconcile_marketplace("Lazada", mp_df, tracker["existing_order_ids"], tc_df)
        results.append(res)
        all_new_rows.extend(res.new_rows)

    if tiktok_file is not None:
        mp_df = read_tiktok_returns(tiktok_file)
        res = reconcile_marketplace("TikTok", mp_df, tracker["existing_order_ids"], tc_df)
        results.append(res)
        all_new_rows.extend(res.new_rows)

    if all_new_rows:
        append_new_rows(tracker, all_new_rows)

    return tracker["workbook"], results
